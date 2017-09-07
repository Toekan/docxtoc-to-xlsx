import docx
import openpyxl as op

import os

from tkinter import filedialog
from tkinter import *

sys.path.insert(0, os.getcwd())

import docxtoc2excel.fromdocx.toc_parser as toc_parser
import docxtoc2excel.toexcel.styles as styles
import docxtoc2excel.toexcel.match_apply as match_apply


# Open browse frame to choose docx file
root = Tk()
root.filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                           filetypes=(("docx files", "*.docx"),
                                                      ("all files", "*.*")))
path_docx = root.filename

# Chosen content of the top row in excel, so the names of each column
firstrow = ['art. Nr.', '', 'omschrijving', 'aantal', 'factor',
            'lengte (y)', 'breedte (x)', 'hoogte (z)', 'hoeveelheid',
            'EENHEIDSPRIJS', 'TOTAAL']

# Excel file saved in same folder with same name
path_excel = re.sub(r'.docx', r'.xlsx', path_docx)

# Load docx file
doc = docx.Document(path_docx)

# Create Excel sheet
wb = op.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

# Inserting top row with lay-out
for cell, title in zip(list(sheet.iter_rows('A1:K1'))[0], firstrow):
    cell.value = title
    cell.font = styles.font_titles
    cell.alignment = styles.alignment_titles
    cell.border = styles.thick_bottom_border

# Column and row dimensions
sheet.row_dimensions[1].height = 20
sheet.column_dimensions['C'].width = 100
sheet.column_dimensions['J'].width = 40
sheet.column_dimensions['K'].width = 40


list_tokenized_toc = toc_parser.import_and_tokenize_toc(doc)

list_match_apply = [(match_apply.match_row_with_sum, match_apply.apply_row_with_sum),
                    (match_apply.match_row_bold, match_apply.apply_row_bold),
                    (match_apply.match_row_default, match_apply.apply_row_default)]


counter = 0
generator_through_sheet = sheet.iter_rows('A3:K10000')

for excel_row in generator_through_sheet:
    # Goes through the list of lines of the table of content
    # and creates the correct excel worksheet.

    # Take next row of the tokenized table of content
    if counter < len(list_tokenized_toc):
        docx_line = list_tokenized_toc[counter][:]
    else:
        break

    for match, apply in list_match_apply:
        if match(docx_line):
            apply(excel_row, docx_line, generator_through_sheet)
            break

    counter += 1

wb.save(path_excel)

